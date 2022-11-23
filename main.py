import csv
import math
import re
import matplotlib.pyplot as plt
import pdfkit
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from jinja2 import Environment, FileSystemLoader
from prettytable import prettytable

"""Названия столбцов файла"""
titles = ['Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия', 'Компания', 'Оклад', 'Название региона',
          'Дата публикации вакансии']

"""Рейтинг опыта работы"""
experience_rate = {"noExperience": 0, "between1And3": 1, "between3And6": 2, "moreThan6": 3}


class DataSet:
    """Класс для обработки csv файла

    Attributes:
        is_empty (bool): Показывает, пустой ли файл
        error_massage (str): Сообщение об ошибке
        file_name (str): Имя обрабатываемого файла
        vacancies_objects (list): Объекты вакансий
    """

    def __init__(self, file_name):
        """Инициализирует объект DataSet, парсит csv файл

        Args:
            file_name (str): Имя обрабатываемого файла
        """
        self.is_empty = False
        self.error_massage = ''
        self.file_name = file_name
        self.vacancies_objects = self.csv_reader()

    filter_parameters = {
        '': lambda vacancies, some_param: vacancies,
        "Навыки": lambda vacancies, skills: filter(lambda v: all(s in v.key_skills for s in skills), vacancies),
        "Оклад": lambda vacancies, salary:
        filter(
            lambda v: math.floor(v.salary.salary_from) <= math.floor(float(salary)) <= math.floor(v.salary.salary_to),
            vacancies),
        "Дата публикации вакансии": lambda vacancies, date:
        filter(lambda v: f'{v.published_at[8:10]}.{v.published_at[5:7]}.{v.published_at[0:4]}' == date, vacancies),
        "Опыт работы": lambda vacancies, experience:
        filter(lambda v: v.experience_id == v.reversedTranslate[experience], vacancies),
        "Премиум-вакансия": lambda vacancies, premium: filter(lambda v: v.premium == v.reverse_premium(premium),
                                                              vacancies),
        "Идентификатор валюты оклада": lambda vacancies, currency:
        filter(lambda v: v.salary.salary_currency == v.reversedTranslate[currency], vacancies),
        "Название": lambda vacancies, name: filter(lambda v: v.name == name, vacancies),
        "Название региона": lambda vacancies, area: filter(lambda v: v.area_name == area, vacancies),
        "Компания": lambda vacancies, employer_name: filter(lambda v: v.employer_name == employer_name, vacancies)
    }

    sort_parameters = {
        'Название': lambda v: v.name,
        'Описание': lambda v: v.description,
        'Навыки': lambda v: len(v.key_skills),
        'Опыт работы': lambda v: experience_rate[v.experience_id],
        'Премиум-вакансия': lambda v: v.premium,
        'Компания': lambda v: v.employer_name,
        'Оклад': lambda v: v.salary.average_salary,
        'Название региона': lambda v: v.area_name,
        'Дата публикации вакансии': lambda v:
        (f'{v.published_at[0:4]}.{v.published_at[5:7]}.{v.published_at[8:10]}',
         f'{v.published_at[11:13]}.{v.published_at[14:16]}.{v.published_at[17:19]}')
    }

    def csv_reader(self) -> tuple or str:
        """Проходится по csv файлу и парсит данные

        :return: list or str: Возвращает список вакансий либо строку в случае ошибки
        """
        is_header = True
        vacancies = []
        header = []
        header_length = 0
        with open(self.file_name, encoding='utf-8') as csv_file:
            file = csv.reader(csv_file)
            for row in file:
                if is_header:
                    is_header = False
                    header = row
                    header_length = len(header)
                    continue
                if self.is_correct_line(row, header_length):
                    vacancies.append(row)
        if len(header) == 0:
            self.is_empty = True
            self.error_massage = 'Пустой файл'
        elif len(vacancies) == 0:
            self.is_empty = True
            self.error_massage = 'Нет данных'
        return self.formatter(header, vacancies)

    def is_correct_line(self, line: list, header_length: int) -> bool:
        """Проверяет строку на пустые ячейки и наличие всех ячеек

        Args:
            line (list): Строка для проверки
            header_length (int): Необходимое количество ячеек
        Returns:
             bool: Значение, корректна ли строка
        """
        return '' not in line and len(line) == header_length

    def clean_line(self, line: str) -> str:
        """Чистит строку от лишних пробелов и html тэгов

        Args:
            line (str): Строка для очистки
        Returns:
            str: Очищенная строка
        """
        line = re.sub('<[^<]+?>', '', line).replace('\xa0', ' ').replace(" ", ' ').strip()
        while '  ' in line:
            line = line.replace('  ', ' ')
        return line

    def formatter(self, header: list, rows: list) -> list:
        """Очищает все строки csv таблицы и превращает их в объекты вакансий

        Args:
            header (list): Список заголовков
            rows (list): Список строк
        Returns:
             list: Список вакансий
        """
        vacancies = []
        for row in rows:
            vacancies_dict = {}
            for i in range(len(header)):
                vacancies_dict[header[i]] = self.clean_line(row[i])
            vacancies.append(Vacancy(vacancies_dict))
        return vacancies

    def filter_sort_data(self, filter_key: str, filter_value: str, sort_param: str, is_rev):
        """Фильтрует и сортирует вакансии по заданным параметрам

        Args:
            filter_key (str): Ключ, по которому будет произведена фильтрация (Название столбца)
            filter_value (str): Значение, по которому будет происходить фильтрация
            sort_param (str): Название столбца по которому будет произведена сортировка
            is_rev (bool): Значение, определяющее порядок сортировки
        """
        self.vacancies_objects = list(self.filter_parameters[filter_key](self.vacancies_objects, filter_value))
        self.vacancies_objects = self.vacancies_objects if sort_param == '' \
            else sorted(self.vacancies_objects, key=self.sort_parameters[sort_param], reverse=is_rev)


class Vacancy:
    """Класс для представления вакансий

    Attributes:
        name (str): Имя вакансии
        description (str): Описание вакансии
        key_skills (list): Необходимые навыки
        experience_id (str): Необходимый опыт работы
        premium (str): Является ли вакансия премиум
        employer_name (str): Имя работодателя
        salary (Salary): Информация о зарплате
        area_name (str): Место работы
        published_at (str): Время публикации
    """

    def __init__(self, args: dict):
        """Инициализирует вакансию

        Args:
            args (dict): Словарь для всех полей вакансии
        """
        self.name = args['name'] if 'name' in args.keys() else ''
        self.description = args['description'] if 'description' in args.keys() else ''
        self.key_skills = args['key_skills'].split('\n') if 'key_skills' in args.keys() else ''
        self.experience_id = args['experience_id'] if 'experience_id' in args.keys() else ''
        self.premium = args['premium'] if 'premium' in args.keys() else ''
        self.employer_name = args['employer_name'] if 'employer_name' in args.keys() else ''
        self.salary = Salary(args)
        self.area_name = args['area_name'] if 'area_name' in args.keys() else ''
        self.published_at = args['published_at'] if 'published_at' in args.keys() else ''

    translate = {"noExperience": "Нет опыта", "between1And3": "От 1 года до 3 лет", "between3And6": "От 3 до 6 лет",
                 "moreThan6": "Более 6 лет", "AZN": "Манаты", "BYR": "Белорусские рубли", "EUR": "Евро",
                 "GEL": "Грузинский лари", "KGS": "Киргизский сом", "KZT": "Тенге", "RUR": "Рубли", "UAH": "Гривны",
                 "USD": "Доллары", "UZS": "Узбекский сум", "TRUE": 'Без вычета налогов', "True": 'Без вычета налогов',
                 "FALSE": 'С вычетом налогов', "False": 'С вычетом налогов'
                 }
    reversedTranslate = dict(zip(translate.values(), translate.keys()))

    def get_published_at_year(self) -> int:
        """Возвращает год публикации

        Returns:
            int: Год публикации
        """
        return int(self.published_at[0:4])

    def get_russian_format(self) -> list:
        """Возвращает список из полей вакансии в русском формате

        Returns:
            list: Список форматированных полей
        """
        return [self.name, self.cut_line(self.description), self.cut_line('\n'.join(self.key_skills)),
                self.translate[self.experience_id], self.premium_yes_no(), self.employer_name,
                f'{self.format_number(self.salary.salary_from)} - {self.format_number(self.salary.salary_to)}'
                f' ({self.translate[self.salary.salary_currency]}) ({self.translate[self.salary.salary_gross]})',
                self.area_name, f'{self.published_at[8:10]}.{self.published_at[5:7]}.{self.published_at[0:4]}']

    def premium_yes_no(self) -> str:
        """Возвращает Да/Нет в зависимости от булева значения premium

        Returns:
            str: Да/Нет
        """
        return 'Да' if self.premium == 'True' else 'Нет'

    def cut_line(self, line: str) -> str:
        """Обрезает линию до 100 символов и добавляет в конце многоточие

        Args:
            line (str): Линия для обрезки
        Returns:
            str: Обрезанная линия
        """
        return line[0:100] + '...' if len(line) > 100 else line

    def format_number(self, number: float) -> str:
        """Задает формат числу вида "ddd ddd"

        Args:
            number (int): Число для форматирования
        Returns:
            int: Отформатированное число
        """
        return '{:3,d}'.format(math.floor(float(number))).replace(',', ' ')

    def reverse_premium(self, premium: str) -> str:
        """Возвращает булево значение в зависимости от вхождения Да/Нет

        Args:
            premium (str): Да/Нет
        Returns:
             str: строковое представление булева значения
        """
        return 'True' if premium == 'Да' else 'False'


class Salary:
    """Класс для представления зарплаты

    Attributes:
        salary_from (float): Нижняя граница оклада
        salary_to (float): Верхняя граница оклада
        salary_gross (str): Строковое представления, указан ли оклад до вычета налогов
        salary_currency (str): Валюта оклада
    """

    def __init__(self, args: dict):
        """Инициальзирует класс, выполняет конвертацию

        Args:
            args (dict): Словарь с ключами для каждого поля класса
        """
        self.salary_from = float(args['salary_from']) if 'salary_from' in args.keys() else 0.0
        self.salary_to = float(args['salary_to']) if 'salary_to' in args.keys() else 0.0
        self.salary_gross = args['salary_gross'] if 'salary_gross' in args.keys() else 0.0
        self.salary_currency = args['salary_currency'] if 'salary_currency' in args.keys() else 0.0
        self.average_salary = math.floor(
            ((self.salary_from + self.salary_to) / 2) * self.currency_to_rub[self.salary_currency])

    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }


class InputConnect:
    """Класс для обработки ввода и вызова дальнейших инструкций, в зависимости от ввода (точка входа)

    Attributes:
        function_selection (str): Выбор программы (Вакансии/Статистика)
        is_input_correct (bool): Является ли ввод корректным
        error_message (str): Сообщение об ошибке при некорректном вводе
    """

    def __init__(self):
        """Запрашивает данные из консоли, и, в зависимости от ввода, выбирает дальнейшие инструкции"""
        self.function_selection = input('Что необходимо сделать (Вакансии или Статистика): ')
        if self.function_selection == 'Вакансии':
            self.error_message = ''
            self.is_input_correct = True
            if not self.is_input_correct:
                print(self.error_message)
                return
            self.print_vacancies()
        elif self.function_selection == 'Статистика':
            self.get_report()
        else:
            print('Неверный ввод данных')

    def print_vacancies(self) -> None:
        """Собирает параметры печати таблица и вызывает печать"""
        file_name = input('Введите название файла: ')
        filter_key, filter_value = self.get_filter_parameter()
        sort_parameter = self.get_sort_parameter()
        sort_reversed = self.get_sort_reversed()
        range_to_print = self.get_range_to_print()
        titles_to_print = self.get_titles_to_print()

        self.print_table(file_name, filter_key, filter_value, sort_parameter, sort_reversed, range_to_print,
                         titles_to_print)

    def get_report(self) -> None:
        """Собирает параметры отчета и вызывает формирование отчета"""
        file_name = input('Введите название файла: ')
        vacancy_name = input('Введите название профессии: ')
        Report(file_name, vacancy_name)

    def get_filter_parameter(self) -> tuple:
        """Получает и обрабатывает параметр фильтрации

        Returns:
            tuple: Ключ и значение для фильтрации
        """
        filter_parameter = input('Введите параметр фильтрации: ')
        key, value = '', ''
        if ': ' in filter_parameter:
            key, value = filter_parameter.split(': ')
            if key == "Навыки":
                value = value.split(', ')

        if filter_parameter != '' and ': ' not in filter_parameter:
            self.error_message = "Формат ввода некорректен"
            self.is_input_correct = False

        return key, value

    def get_sort_parameter(self) -> str:
        """Получает и обрабатывает параметр сортировки

        Returns:
            str: Параметр сортировки
        """
        sort_parameter = input('Введите параметр сортировки: ')
        if sort_parameter not in titles and sort_parameter != '':
            self.error_message = "Параметр сортировки некорректен" if self.error_message == '' else self.error_message
            self.is_input_correct = False
        return sort_parameter

    def get_sort_reversed(self) -> bool:
        """Получает и обрабатывает направление сортировки

        Returns:
            bool: Значение, обратный ли порядок
        """
        sort_direction = input('Обратный порядок сортировки (Да / Нет): ')
        if sort_direction == 'Да':
            sort_direction = True
        elif sort_direction == 'Нет' or sort_direction == '':
            sort_direction = False
        if type(sort_direction) != bool:
            self.error_message = "Порядок сортировки задан некорректно" if self.error_message == '' else self.error_message
            self.is_input_correct = False
        return sort_direction

    def get_range_to_print(self) -> tuple:
        """Получает и обрабатывает диапазон

        Returns:
            tuple: Первый и последний элемент диапазона
        """
        start = 0
        end = 100000000000
        range_to_print = input('Введите диапазон вывода: ').split()
        if len(range_to_print) == 2:
            start, end = int(range_to_print[0]) - 1, int(range_to_print[1]) - 1
        elif len(range_to_print) == 1:
            start = int(range_to_print[0]) - 1
        return start, end

    def get_titles_to_print(self) -> list:
        """Получает и обрабатывает заголовки, которые необходимо напечатать

        Returns:
            list: Заголовки, которые необходимо напечатать
        """
        titles_to_print = input('Введите требуемые столбцы: ')
        return titles_to_print.split(', ') if len(titles_to_print) > 1 else [titles_to_print]

    def print_table(self, file_name: str, filter_key: str, filter_value: str, sort_parameter: str, sort_reversed: bool,
                    range_to_print: tuple, titles_to_print: list) -> None:
        """Печатает таблицу с вакансиями в консоль

        Args:
            file_name (str): Имя файла
            filter_key (str): Ключ фильтрации
            filter_value (str): Значение фильтрации
            sort_parameter (str): Параметр сортировки
            sort_reversed (bool): Является ли сортировка обратной
            range_to_print (tuple): Диапазон печати
            titles_to_print (list): Столбцы для печати
        """
        data_set = DataSet(file_name)
        if data_set.is_empty:
            print(data_set.error_massage)
            return
        data_set.filter_sort_data(filter_key, filter_value, sort_parameter, sort_reversed)
        if len(data_set.vacancies_objects) == 0:
            print('Ничего не найдено')
            return

        counter = 0
        table = prettytable.PrettyTable(
            hrules=prettytable.ALL,
            align='l',
            field_names=['№'] + titles,
            max_width=20)
        for vacancie in data_set.vacancies_objects:
            counter += 1
            table.add_row([counter] + vacancie.get_russian_format())
        print(table.get_string(start=range_to_print[0],
                               end=range_to_print[1],
                               fields=['№'] + titles_to_print if len(titles_to_print) != 0 else [
                                                                                                    '№'] + table.field_names))


class Statistics:
    """Класс для сбора статистики по вакансиям

    Attributes:
        vacancies (list): Список вакансий
        vacancie_name (str): Название вакансии, для которой собирается статистика
        salary_by_year (dict): Статистика зарплат по годам
        count_by_year (dict): Статистика количества вакансий по годам
        prof_salary_by_year (dict): Статистика зарплат по годам для заданной профессии
        prof_count_by_year (dict):Статистика количества вакансий по годам для заданной профессии
        salary_by_city (dict): Статистика зарплат по городам
        count_by_city (dict): Статистика количества вакансий по городам
    """

    def __init__(self, file_name: str, vacancie_name: str):
        """Инициализирует класс, получает данные по вакансиям

        Args:
            file_name (str): Имя csv файла, откуда взять вакансии
        """
        self.vacancies = DataSet(file_name).vacancies_objects
        self.vacancie_name = vacancie_name
        self.salary_by_year = {}
        self.count_by_year = {}
        self.prof_salary_by_year = {}
        self.prof_count_by_year = {}
        self.salary_by_city = {}
        self.count_by_city = {}

    def get_statistics(self) -> None:
        """Собирает статистику по вакансиям и выводит ее в консоль"""
        for vacancy in self.vacancies:
            self.update_stats(vacancy.get_published_at_year(), vacancy.salary.average_salary, self.salary_by_year)
            self.update_stats(vacancy.get_published_at_year(), 1, self.count_by_year)
            self.update_stats(vacancy.area_name, vacancy.salary.average_salary, self.salary_by_city)
            self.update_stats(vacancy.area_name, 1, self.count_by_city)
            if self.vacancie_name in vacancy.name:
                self.update_stats(vacancy.get_published_at_year(), vacancy.salary.average_salary,
                                  self.prof_salary_by_year)
                self.update_stats(vacancy.get_published_at_year(), 1, self.prof_count_by_year)

        self.get_average_salary(self.salary_by_year, self.count_by_year)
        self.get_average_salary(self.prof_salary_by_year, self.prof_count_by_year)
        self.get_average_salary(self.salary_by_city, self.count_by_city)
        self.get_percentage_of_total(self.count_by_city, len(self.vacancies))

        self.get_cities_with_enough_count()
        self.sort_statistics()
        self.print_stats()

    def update_stats(self, key: str, value: int or float, stats: dict) -> None:
        """Обновляет статистику для новой вакансии

        Args:
            key (str): Ключ, для которого необходимо обновить статистику
            value (int or float): На сколько необходимо увеличить статистику
            stats (dict): Для какой статистики необходимо обновить данные
        """
        if key in stats.keys():
            stats[key] += value
        else:
            stats[key] = value

    def get_average_salary(self, salary_stats: dict, count_stats: dict) -> None:
        """Считает среднюю зарплату для каждого ключа статистики

        Args:
            salary_stats (dict): Статистика зарплат
            count_stats (dict): Статистика количества вакансий
        """
        for key in count_stats.keys():
            salary_stats[key] = math.floor(salary_stats[key] / count_stats[key])

    def get_percentage_of_total(self, stats: dict, count: int) -> None:
        """Считает процент количества вакансий по городам от общего количества

        Args:
            stats (dict): Статистика для изменения
            count (int): Общее количество вакансий
        """
        for key in stats.keys():
            stats[key] = round(stats[key] / count, 4)

    def get_cities_with_enough_count(self) -> None:
        """Отбирает города с достаточным количеством вакансий"""
        new_salary_by_city = {}
        new_count_by_city = {}
        for key, value in self.count_by_city.items():
            if value * 100 < 1:
                continue
            new_salary_by_city[key] = self.salary_by_city[key]
            new_count_by_city[key] = value
        self.salary_by_city = new_salary_by_city
        self.count_by_city = new_count_by_city

    def sort_statistics(self) -> None:
        """Сортирует статистику"""
        self.prof_salary_by_year = self.prof_salary_by_year if len(self.prof_salary_by_year) != 0 else {2022: 0}
        self.prof_count_by_year = self.prof_count_by_year if len(self.prof_count_by_year) != 0 else {2022: 0}
        self.salary_by_city = dict(sorted(self.salary_by_city.items(), key=lambda x: x[1], reverse=True)[:10])
        self.count_by_city = dict(sorted(self.count_by_city.items(), key=lambda x: x[1], reverse=True)[:10])

    def get_percent_of_other_cities(self) -> None:
        """Считает процент городов, не вошедших в статистику"""
        others_percent = 1
        for city_percent in self.count_by_city.values():
            others_percent -= city_percent
        self.count_by_city['Другие'] = others_percent

    def print_stats(self) -> None:
        """Выводит в консоль статистику"""
        print('Динамика уровня зарплат по годам:', self.salary_by_year)
        print('Динамика количества вакансий по годам:', self.count_by_year)
        print('Динамика уровня зарплат по годам для выбранной профессии:', self.prof_salary_by_year)
        print('Динамика количества вакансий по годам для выбранной профессии:', self.prof_count_by_year)
        print('Уровень зарплат по городам (в порядке убывания):', self.salary_by_city)
        print('Доля вакансий по городам (в порядке убывания):', self.count_by_city)


class GraphsCreator:
    """Класс для создания диаграмм по статистике"""

    def __init__(self, stats):
        """Генерирует изображение c диаграммами

        Args:
            stats (Statistics): Статистика по вакансиям
        """
        self.generate_image(stats)

    def generate_image(self, stats: Statistics) -> None:
        """Генерирует 4 диаграммы и сохраняет их в изображение

        Args:
            stats (Statistics): Статистика по вакансиям
        """
        plt.subplots(figsize=(10, 7))
        plt.grid(True)

        self.create_salary_by_year_plot(stats)
        self.create_count_by_year_plot(stats)
        self.create_salary_by_city_plot(stats)
        self.create_count_by_city_plot(stats)

        plt.subplots_adjust(wspace=0.5, hspace=0.5)
        plt.savefig('graph.png')

    def create_salary_by_year_plot(self, stats: Statistics) -> None:
        """Генерирует столбчатую диаграмму по уровню зарплат по годам

        Args:
            stats (Statistics): Статистика по вакансиям
        """
        first = plt.subplot(221)
        plt.tick_params(axis='x', which='major', labelsize=8, rotation=90)
        plt.tick_params(axis='y', which='major', labelsize=8)
        first.bar(list(map(lambda y: y - 0.2, stats.salary_by_year.keys())),
                  stats.salary_by_year.values(), width=0.4,
                  label='Средняя з/п')
        first.bar(list(map(lambda y: y + 0.2, stats.prof_salary_by_year.keys())),
                  stats.prof_salary_by_year.values(), width=0.4,
                  label=f'З/п {stats.vacancie_name}')
        plt.legend(fontsize=8)
        plt.title('Уровень зарплат по годам', fontsize=12)

    def create_count_by_year_plot(self, stats: Statistics) -> None:
        """Генерирует столбчатую диаграмму по количеству вакансий по годам

        Args:
            stats (Statistics): Статистика по вакансиям
        """
        second = plt.subplot(222)
        plt.tick_params(axis='x', which='major', labelsize=8, rotation=90)
        plt.tick_params(axis='y', which='major', labelsize=8)
        second.bar(list(map(lambda y: y - 0.2, stats.count_by_year.keys())),
                   stats.count_by_year.values(), width=0.4,
                   label='Количество вакансий')
        second.bar(list(map(lambda y: y + 0.2, stats.prof_count_by_year.keys())),
                   stats.prof_count_by_year.values(), width=0.4,
                   label=f'Количество вакансий {stats.vacancie_name}')
        plt.legend(fontsize=8)
        plt.title('Количество вакансий по годам', fontsize=12)

    def create_salary_by_city_plot(self, stats: Statistics) -> None:
        """Генерирует горизонтальную столбчатую диаграмму по уровню зарплат по городам

        Args:
            stats (Statistics): Статистика по вакансиям
        """
        third = plt.subplot(223)
        plt.tick_params(axis='x', which='major', labelsize=8)
        plt.tick_params(axis='y', which='major', labelsize=6)
        third.barh(list(reversed(stats.salary_by_city.keys())),
                   list(reversed(stats.salary_by_city.values())))
        plt.title('Уровень зарплат по городам', fontsize=12)

    def create_count_by_city_plot(self, stats: Statistics) -> None:
        """Генерирует круговую диаграмму по количеству вакансий по городам

        Args:
            stats (Statistics): Статистика по вакансиям
        """
        stats.get_percent_of_other_cities()
        fourth = plt.subplot(224)
        plt.rc('xtick', labelsize=6)
        fourth.pie(list(map(lambda c: round(c * 100, 2), stats.count_by_city.values())),
                   labels=stats.count_by_city.keys(),
                   colors=['r', 'g', 'b', 'm', 'y', 'c', 'orange', 'darkblue', 'pink', 'sienna', 'grey'])
        plt.title('Доля вакансий по городам', fontsize=12)


class ExcelCreator:
    """Класс для создания Excel таблицы со статистикой

    Attributes:
        workbook (Workbook): Excel таблица
    """

    def __init__(self, stats: Statistics):
        """Инициализирует класс, создавая таблицу

        Args:
            stats (Statistics): Статистика по вакансиям
        """
        self.workbook = self.initialize_workbook(stats)

    def initialize_workbook(self, stats: Statistics) -> Workbook:
        """Инициализирует таблицу и добавляет в нее данные

        Args:
            stats (Statistics): Статистика по вакансиям
        Returns:
            Workbook: excel таблица
        """
        workbook, years_sheet, cities_sheet = self.create_workbook(stats.vacancie_name)
        self.add_stats_to_excel(stats, years_sheet, cities_sheet)
        self.set_sheets_settings(stats, years_sheet, cities_sheet)
        workbook.save('report.xlsx')
        return workbook

    def create_workbook(self, vacancie_name: str) -> tuple:
        """Создает excel таблицу, добавляет в нее 2 листа, задает заголовки

        Args:
            vacancie_name (str): Название профессии, по которой собрана статистика
        Returns:
            tuple: Excel таблица, 2 листа таблицы
        """
        workbook = Workbook()

        years_sheet = workbook.active
        years_sheet.title = 'Статистика по годам'
        cities_sheet = workbook.create_sheet('Статистика по городам')

        years_sheet.append(
            ['Год', 'Средняя зарплата', f'Средняя зарплата - {vacancie_name}',
             'Количество вакансий', f'Количество вакансий - {vacancie_name}'])
        cities_sheet.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля Вакансий'])

        return workbook, years_sheet, cities_sheet

    def add_stats_to_excel(self, stats: Statistics, years_sheet: Workbook.worksheets,
                           cities_sheet: Workbook.worksheets):
        """Добавляет статистику в таблицу excel

        Args:
            stats (Statistics): Статистика по вакансиям
            years_sheet (Workbook.worksheets): Excel лист с статистикой по годам
            cities_sheet (Workbook.worksheets): Excel лист с статистикой по городам
        """
        for year in stats.salary_by_year.keys():
            years_sheet.append([year,
                                stats.salary_by_year[year],
                                stats.prof_salary_by_year[year],
                                stats.count_by_year[year],
                                stats.prof_salary_by_year[year]])

        for city in stats.salary_by_city.keys():
            cities_sheet.append([city, stats.salary_by_city[city]])

        for i, city in enumerate(stats.count_by_city.keys(), 2):
            cities_sheet[f'D{i}'].value = city
            cities_sheet[f'E{i}'].value = f'{round(stats.count_by_city[city] * 100, 2)}%'

    def set_sheets_settings(self, stats: Statistics, years_sheet: Workbook.worksheets,
                            cities_sheet: Workbook.worksheets) -> None:
        """Устанавливает настройки для таблицы

        Args:
            stats (Statistics): Статистика по вакансиям
            years_sheet (Workbook.worksheets): Excel лист с статистикой по годам
            cities_sheet (Workbook.worksheets): Excel лист с статистикой по городам
        """
        used_columns = ['A', 'B', 'C', 'D', 'E']
        for i in used_columns:
            years_sheet[f'{i}1'].font = Font(bold=True)
            cities_sheet[f'{i}1'].font = Font(bold=True)
            years_sheet.column_dimensions[i].width = max(map(lambda x: len(str(x.value)), years_sheet[i])) + 1
            cities_sheet.column_dimensions[i].width = max(
                map(lambda x: len(str(x.value)), cities_sheet[i])) + 1

        thins = Side(border_style="thin")
        for column in used_columns:
            for row in range(1, len(stats.salary_by_year.keys()) + 2):
                years_sheet[f'{column}{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)

        for column in used_columns:
            for row in range(1, len(stats.salary_by_city.keys()) + 2):
                if column == 'C':
                    break
                cities_sheet[f'{column}{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)


class PdfCreator:
    """Класс для создания отчета в виде pdf файла"""

    def __init__(self, vacancie_name: str, workbook: Workbook, years_sheet_rows: int, cities_sheet_rows: int):
        """Задает настройки и создает pdf отчет

        Args:
            vacancie_name (str): Название вакансии, по которой была собрана статистика
            workbook (Workbook): Таблица excel
            years_sheet_rows (int): Количество строк в таблице по годам
            cities_sheet_rows (int): Количество строк в таблице по городам
        """
        config = pdfkit.configuration(wkhtmltopdf=r'D:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(self.fill_pdf_template(vacancie_name,
                                                  workbook["Статистика по годам"],
                                                  workbook["Статистика по городам"],
                                                  years_sheet_rows,
                                                  cities_sheet_rows),
                           'report.pdf', configuration=config, options=options)

    def fill_pdf_template(self, vacancie_name: str, years_sheet: Workbook.worksheets, cities_sheet: Workbook.worksheets,
                          years_sheet_rows: int, cities_sheet_rows: int) -> str:
        """Заполняет html template данными из таблицы

        Args:
            vacancie_name (str): Название вакансии, по которой была собрана статистика
            years_sheet (Workbook.worksheets): Excel лист с статистикой по годам
            cities_sheet (Workbook.worksheets): Excel лист с статистикой по городам
            years_sheet_rows (int): Количество строк в таблице по годам
            cities_sheet_rows (int): Количество строк в таблице по городам
        Returns:
            str: Заполненный данными из таблицы html template
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('pdf_template.html')
        pdf_template = template.render({'vacancie_name': vacancie_name,
                                        'years_table': self.create_html_table(years_sheet, years_sheet_rows),
                                        'cities_table_first': self.create_html_table(cities_sheet,
                                                                                     cities_sheet_rows, last_column=2),
                                        'cities_table_second': self.create_html_table(cities_sheet,
                                                                                      cities_sheet_rows, 4)})
        return pdf_template

    def create_html_table(self, ws: Workbook.worksheets, rows_count: int, first_column: int = 1,
                          last_column: int = 5) -> str:
        """Создает html таблицу, заполненную данными из ws

        Args:
            ws: Workbook.worksheets: Лист таблицы excel
            rows_count (int): Количество строк
            first_column (int): Первая колонка
            last_column (int): Последняя колонка
        Returns:
            str: Строка html таблицы
        """
        html = ''
        is_first = True
        for row in ws.iter_rows(min_row=1, min_col=first_column, max_col=last_column, max_row=rows_count + 1):
            html += '<tr>'
            for cell in row:
                html += '<td><b>' + str(cell.value) + '</b></td>' if is_first else '<td>' + str(cell.value) + '</td>'
            html += '</tr>'
            is_first = False
        return html


class Report:
    """Класс для создания отчета: excel таблицы, изображения с диаграммами и общего отчета в pdf

    Attributes:
        statistics (Statistics): Статистика по вакансиям
    """

    def __init__(self, file_name: str, vacancie_name: str):
        """Собирает статистику и создает отчет

        Args:
            file_name (str): Имя файла с данными
            vacancie_name (str): Название вакансии, по которой собирается статистика
        """
        self.statistics = Statistics(file_name, vacancie_name)
        self.statistics.get_statistics()

        GraphsCreator(self.statistics)
        PdfCreator(vacancie_name, ExcelCreator(self.statistics).workbook, len(self.statistics.salary_by_year.keys()),
                   len(self.statistics.salary_by_city.keys()))


InputConnect()
